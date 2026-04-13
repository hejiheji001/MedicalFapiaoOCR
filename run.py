#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MedicalFapiaoOCR — 医疗费发票OCR汇总工具

从就医记录PDF中提取所有医疗费发票，按发票号码汇总金额和页码，输出Excel汇总表。
支持电子票据文本提取和扫描件OCR识别，并行处理。

Usage:
    python run.py <pdf_path> [-o output_dir] [-w workers] [--dpi 300] [--save-raw]
"""

import argparse
import io
import json
import os
import re
import sys
import time
from collections import OrderedDict
from concurrent.futures import ProcessPoolExecutor, as_completed

import fitz
import numpy as np
from PIL import Image


# ============================================================
# PDF page image extraction
# ============================================================

def extract_page_image(pdf_path, page_idx, dpi=300):
    """Extract a single page as RGB PNG bytes."""
    doc = fitz.open(pdf_path)
    page = doc[page_idx]
    pix = page.get_pixmap(dpi=dpi, colorspace=fitz.csRGB)
    img_bytes = pix.tobytes("png")
    doc.close()
    return img_bytes


# ============================================================
# OCR for scanned pages (runs in subprocess)
# ============================================================

def ocr_batch_pages(args):
    """OCR a batch of scanned pages.

    Runs inside a subprocess via ProcessPoolExecutor.
    The OCR engine is initialised once per worker to amortise startup cost.
    """
    page_indices, pdf_path, dpi = args

    from rapidocr_onnxruntime import RapidOCR
    ocr = RapidOCR()

    results = []
    for page_idx in page_indices:
        page_num = page_idx + 1
        t0 = time.time()

        img_bytes = extract_page_image(pdf_path, page_idx, dpi)
        img = Image.open(io.BytesIO(img_bytes))
        img_np = np.array(img)

        result, _ = ocr(img_np)
        elapsed = time.time() - t0

        if result is None:
            results.append({
                'page': page_num,
                'type': 'NO_TEXT',
                'invoice_num': None,
                'amount': None,
                'elapsed': elapsed,
            })
            continue

        lines = [item[1] for item in result]
        full_text = ' '.join(lines)

        invoice_num = None
        amount = None
        page_type = 'UNKNOWN'

        # --- invoice number ---
        for line in lines:
            m = re.search(r'票据号码[：:]\s*(\d{5,})', line)
            if m:
                invoice_num = m.group(1)
                break
            m = re.search(r'号码[：:]\s*(\d{5,})', line)
            if m:
                invoice_num = m.group(1)
                break

        # --- associated invoice (detail pages) ---
        assoc_invoice = None
        for line in lines:
            if '所属' in line or '所' in line:
                m = re.search(r'(\d{8,})', line)
                if m:
                    assoc_invoice = m.group(1)
                    break

        # --- page type: RECEIPT if amount found ---
        for line in lines:
            m = re.search(r'[（(]?\s*小写\s*[)）]?\s*([\d,.]+)', line)
            if m:
                amount = m.group(1).replace(',', '')
                page_type = 'RECEIPT'
                break

        # --- page type: DETAIL if no amount ---
        if page_type != 'RECEIPT':
            detail_text = re.sub(r'含明细', '', full_text)
            if '收费明细' in detail_text or '明细' in detail_text:
                page_type = 'DETAIL'
                if not invoice_num:
                    invoice_num = assoc_invoice

        if page_type == 'UNKNOWN' and any('小计' in l or '合计' in l for l in lines):
            page_type = 'DETAIL'
            if not invoice_num:
                invoice_num = assoc_invoice

        results.append({
            'page': page_num,
            'type': page_type,
            'invoice_num': invoice_num or assoc_invoice,
            'amount': amount,
            'elapsed': elapsed,
        })

    return results


# ============================================================
# Text extraction for embedded-text pages
# ============================================================

def extract_text_pages(pdf_path):
    """Extract data from pages that have an embedded text layer.

    Returns:
        (text_results, scan_page_indices)
    """
    doc = fitz.open(pdf_path)
    results = []
    scan_pages = []

    for i in range(doc.page_count):
        page = doc[i]
        text = page.get_text().strip()
        clean = re.sub(r'第\d+页[,，]共\d+页', '', text).strip()

        if len(clean) > 20:
            invoice_num = None
            amount = None
            page_type = 'UNKNOWN'

            m = re.search(r'票据号码[：:]\s*(\d+)', text)
            if m:
                invoice_num = m.group(1)

            m_assoc = re.search(r'所属电子票据号码[：:]\s*(\d+)', text)
            if m_assoc and not invoice_num:
                invoice_num = m_assoc.group(1)

            if '收费明细' in text:
                page_type = 'DETAIL'
                if not invoice_num and m_assoc:
                    invoice_num = m_assoc.group(1)
            elif '收费票据' in text:
                page_type = 'RECEIPT'
                m_amt = re.search(r'[（(]小写[)）]\s*([\d,.]+)', text)
                if m_amt:
                    amount = m_amt.group(1).replace(',', '')

            results.append({
                'page': i + 1,
                'type': page_type,
                'invoice_num': invoice_num,
                'amount': amount,
                'elapsed': 0,
            })
        else:
            scan_pages.append(i)

    doc.close()
    return results, scan_pages


# ============================================================
# Grouping helpers
# ============================================================

def build_page_ranges(pages_sorted):
    """Build a human-readable page range string.

    Examples:
        [1]           -> '1'
        [1,2,3]       -> '1-3'
        [1,2,5,6,7]   -> '1-2,5-7'
    """
    if not pages_sorted:
        return ''
    if len(pages_sorted) == 1:
        return str(pages_sorted[0])

    ranges = []
    start = end = pages_sorted[0]
    for pg in pages_sorted[1:]:
        if pg == end + 1:
            end = pg
        else:
            ranges.append((start, end))
            start = end = pg
    ranges.append((start, end))

    return ','.join(
        str(s) if s == e else f'{s}-{e}'
        for s, e in ranges
    )


def group_by_invoice(all_pages):
    """Group page records by invoice number.

    Missing invoice numbers on DETAIL/UNKNOWN pages are inherited from the
    nearest preceding page that has one (handles manually-assembled PDFs).
    """
    # Fix missing invoice numbers
    for i, p in enumerate(all_pages):
        if p['type'] in ('DETAIL', 'UNKNOWN') and p['invoice_num'] is None:
            for j in range(i - 1, -1, -1):
                if all_pages[j]['invoice_num']:
                    p['invoice_num'] = all_pages[j]['invoice_num']
                    if p['type'] == 'UNKNOWN':
                        p['type'] = 'DETAIL'
                    break

    invoice_map = OrderedDict()
    for p in all_pages:
        inv = p['invoice_num']
        if inv not in invoice_map:
            invoice_map[inv] = {'amount': 0, 'pages': []}
        if p['amount']:
            invoice_map[inv]['amount'] = float(p['amount'])
        invoice_map[inv]['pages'].append(p['page'])

    for data in invoice_map.values():
        data['page_range'] = build_page_ranges(sorted(data['pages']))

    return invoice_map


# ============================================================
# Excel generation
# ============================================================

def generate_excel(invoices, output_path):
    """Generate an Excel summary matching the 3-column-pair template layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '医疗费汇总'

    title_font = Font(name='SimHei', size=14, bold=True)
    header_font = Font(name='SimSun', size=11, bold=True)
    data_font = Font(name='SimSun', size=11)
    center = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    rows = list(invoices.values())
    total = sum(r['amount'] for r in rows)

    # Title
    ws.merge_cells('A1:F1')
    cell = ws['A1']
    cell.value = f'医疗费汇总：{total:.2f} 元'
    cell.font = title_font
    cell.alignment = center

    # 3 column-pairs
    COL_PAIRS = 3
    rows_per_col = (len(rows) + COL_PAIRS - 1) // COL_PAIRS

    # Headers (row 3)
    for ci in range(COL_PAIRS):
        base = ci * 2 + 1
        for label, off in [('金额', 0), ('页码', 1)]:
            c = ws.cell(row=3, column=base + off, value=label)
            c.font = header_font
            c.alignment = center
            c.border = border
            c.fill = header_fill

    for i in range(6):
        ws.column_dimensions[chr(65 + i)].width = 16

    # Data
    for i, row_data in enumerate(rows):
        col_group = i // rows_per_col
        row_in_group = i % rows_per_col
        excel_row = 4 + row_in_group
        base = col_group * 2 + 1

        amt_cell = ws.cell(row=excel_row, column=base, value=row_data['amount'])
        amt_cell.font = data_font
        amt_cell.alignment = center
        amt_cell.border = border
        amt_cell.number_format = '#,##0.00'

        pg_cell = ws.cell(row=excel_row, column=base + 1, value=row_data['page_range'])
        pg_cell.font = data_font
        pg_cell.alignment = center
        pg_cell.border = border

    # Fill remaining cells with borders
    for r in range(4, 4 + rows_per_col):
        for ci in range(COL_PAIRS):
            for c in (ci * 2 + 1, ci * 2 + 2):
                ws.cell(row=r, column=c).border = border

    wb.save(output_path)
    return total


# ============================================================
# Main pipeline
# ============================================================

def run(pdf_path, output_dir, workers=4, dpi=300, save_raw=False):
    """Execute the full extraction pipeline.

    Args:
        pdf_path:   Path to the medical record PDF.
        output_dir: Directory for output files.
        workers:    Number of parallel OCR workers.
        dpi:        DPI for scanned page rendering.
        save_raw:   Whether to save raw_data.json.

    Returns:
        (total_amount, invoice_count)
    """
    os.makedirs(output_dir, exist_ok=True)
    timings = {}
    t_total = time.time()

    print('=' * 60)
    print('MedicalFapiaoOCR')
    print('=' * 60)
    print(f'  PDF:     {pdf_path}')
    print(f'  Output:  {output_dir}')
    print(f'  Workers: {workers}  DPI: {dpi}')

    # --- Step 1: text pages ---
    t0 = time.time()
    print('\n[1/4] Extracting embedded text pages...')
    text_results, scan_page_indices = extract_text_pages(pdf_path)
    timings['text'] = time.time() - t0
    print(f'  {len(text_results)} text pages, {len(scan_page_indices)} scan pages ({timings["text"]:.2f}s)')

    # --- Step 2: OCR scanned pages ---
    t0 = time.time()
    ocr_results = []
    if scan_page_indices:
        print(f'\n[2/4] OCR scanned pages ({workers} workers)...')
        batches = [[] for _ in range(workers)]
        for i, idx in enumerate(scan_page_indices):
            batches[i % workers].append(idx)
        batches = [b for b in batches if b]

        ocr_args = [(batch, pdf_path, dpi) for batch in batches]

        with ProcessPoolExecutor(max_workers=workers) as executor:
            futures = {executor.submit(ocr_batch_pages, a): i for i, a in enumerate(ocr_args)}
            for future in as_completed(futures):
                batch_results = future.result()
                ocr_results.extend(batch_results)
                idx = futures[future]
                avg = sum(r['elapsed'] for r in batch_results) / len(batch_results)
                print(f'  Worker {idx}: {len(batch_results)} pages, avg {avg:.1f}s/page')
    else:
        print('\n[2/4] No scanned pages — skipping OCR.')
    timings['ocr'] = time.time() - t0
    if scan_page_indices:
        print(f'  OCR done: {timings["ocr"]:.2f}s wall-clock')

    # --- Step 3: merge & group ---
    t0 = time.time()
    print('\n[3/4] Grouping by invoice...')
    all_pages = sorted(text_results + ocr_results, key=lambda x: x['page'])
    invoice_map = group_by_invoice(all_pages)
    timings['group'] = time.time() - t0
    print(f'  {len(invoice_map)} invoices ({timings["group"]:.2f}s)')

    for inv, data in invoice_map.items():
        print(f'  {inv} | {data["amount"]:>10.2f} | {data["page_range"]}')

    # --- Step 4: Excel ---
    t0 = time.time()
    print('\n[4/4] Generating Excel...')
    excel_path = os.path.join(output_dir, '医疗费汇总.xlsx')
    total = generate_excel(invoice_map, excel_path)
    timings['excel'] = time.time() - t0
    print(f'  Saved: {excel_path} ({timings["excel"]:.2f}s)')

    if save_raw:
        raw_path = os.path.join(output_dir, 'raw_data.json')
        with open(raw_path, 'w', encoding='utf-8') as f:
            json.dump(all_pages, f, ensure_ascii=False, indent=2, default=str)
        print(f'  Raw data: {raw_path}')

    # --- Summary ---
    timings['total'] = time.time() - t_total
    print('\n' + '=' * 60)
    print(f'DONE — {total:.2f} 元 ({len(invoice_map)} invoices)')
    print('=' * 60)
    print(f'\n  Text extract:  {timings["text"]:>6.2f}s')
    print(f'  OCR parallel:  {timings["ocr"]:>6.2f}s  ({len(scan_page_indices)} pages, {workers} workers)')
    print(f'  Grouping:      {timings["group"]:>6.2f}s')
    print(f'  Excel:         {timings["excel"]:>6.2f}s')
    print(f'  {"─" * 27}')
    print(f'  TOTAL:         {timings["total"]:>6.2f}s')

    return total, len(invoice_map)


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description='MedicalFapiaoOCR — 医疗费发票OCR汇总工具',
    )
    parser.add_argument('pdf_path', help='输入PDF文件路径')
    parser.add_argument('-o', '--output', default=None, help='输出目录（默认: PDF所在目录）')
    parser.add_argument('-w', '--workers', type=int, default=4, help='并行OCR workers数（默认: 4）')
    parser.add_argument('--dpi', type=int, default=300, help='扫描页面渲染DPI（默认: 300）')
    parser.add_argument('--save-raw', action='store_true', help='保存原始数据到raw_data.json')

    args = parser.parse_args()

    pdf_path = os.path.abspath(args.pdf_path)
    if not os.path.isfile(pdf_path):
        print(f'Error: PDF not found: {pdf_path}', file=sys.stderr)
        sys.exit(1)

    output_dir = args.output or os.path.dirname(pdf_path)

    run(pdf_path, output_dir, workers=args.workers, dpi=args.dpi, save_raw=args.save_raw)


if __name__ == '__main__':
    main()
